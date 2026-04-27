"""
Parse effect_matrix_v4.xlsx -> js/weapons-config.js

Reads the Weapons Master sheet row by row, parses the sim_trigger_key
column as semicolon-separated key=value pairs, and outputs a JS file
declaring const WEAPONS_CONFIG = { weapons: [...], interactions: [...] }.

Loaded via <script> tag at Layer 1 in index.html, before game-state.js.

Usage: python scripts/parse_effect_matrix.py
"""
import openpyxl
import json
import sys
import os

XLSX_PATH = os.path.join(os.path.dirname(__file__), '..', 'Context', 'docs', 'effect_matrix_v4.xlsx')
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), '..', 'js', 'weapons-config.js')


def parse_sim_trigger_keys(raw):
    """Parse 'speed_mult=0.05; spawn_rate_mult=0.2; hormuz_lane=closed' into dict."""
    if not raw:
        return {}
    result = {}
    for pair in raw.split(';'):
        pair = pair.strip()
        if '=' not in pair:
            continue
        key, val = pair.split('=', 1)
        key = key.strip()
        val = val.strip()
        # Try to cast to number or bool
        if val.lower() == 'true':
            result[key] = True
        elif val.lower() == 'false':
            result[key] = False
        elif val.lower() in ('closed', 'open', 'disabled', 'active', 'enabled', 'tanker'):
            result[key] = val.lower()
        else:
            try:
                result[key] = int(val)
            except ValueError:
                try:
                    result[key] = float(val)
                except ValueError:
                    result[key] = val
    return result


def to_num(val, default=0):
    """Safely convert cell value to number."""
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    # --- Weapons Master ---
    ws = wb['Weapons Master']
    rows = list(ws.iter_rows(min_row=1, max_row=50, values_only=True))

    # Find header row (contains '#' in first cell)
    header_idx = None
    for i, row in enumerate(rows):
        if row[0] == '#':
            header_idx = i
            break

    if header_idx is None:
        print('ERROR: Could not find header row with "#" in Weapons Master', file=sys.stderr)
        sys.exit(1)

    headers = [str(h).strip() if h else '' for h in rows[header_idx]]
    print(f'Found headers at row {header_idx + 1}: {headers}')

    # Map header names to indices
    col = {}
    for i, h in enumerate(headers):
        col[h] = i

    weapons = []
    for row in rows[header_idx + 1:]:
        weapon_id = row[col['#']]
        if not weapon_id or not str(weapon_id).startswith(('D', 'R')):
            continue

        weapon = {
            'id': str(weapon_id).strip(),
            'name': str(row[col['weapon_name']] or '').strip(),
            'player': str(row[col['player']] or '').strip(),
            'speed': str(row[col['speed']] or '').strip(),
            'prob_delta': to_num(row[col['prob_delta']]),
            'onset_s': to_num(row[col['onset_s']]),
            'decay_per_30s': to_num(row[col['decay_per_30s']]),
            'build_per_30s': to_num(row[col['build_per_30s']]),
            'sim_consequence': str(row[col['sim_consequence']] or '').strip(),
            'counter_weapons': str(row[col['counter_weapons']] or '').strip(),
            'sim_trigger_keys': parse_sim_trigger_keys(
                str(row[col['sim_trigger_key']] or '')
            ),
            'ships_at_t0': to_num(row[col.get('ships_at_t0\n(N=20, no counters)', '')])
        }
        weapons.append(weapon)

    print(f'\nParsed {len(weapons)} weapons:')
    for w in weapons:
        keys = w['sim_trigger_keys']
        key_str = '; '.join(f'{k}={v}' for k, v in keys.items())
        print(f"  {w['id']:3s} | {w['name'][:40]:<40s} | {w['player']:9s} | {w['speed']:4s} | delta={w['prob_delta']:+.0f} | sim_keys: {key_str}")

    # --- Interactions ---
    ws_int = wb['Interactions']
    int_rows = list(ws_int.iter_rows(min_row=1, max_row=30, values_only=True))

    # Find header row
    int_header_idx = None
    for i, row in enumerate(int_rows):
        if row[0] == 'disruptor_id':
            int_header_idx = i
            break

    interactions = []
    if int_header_idx is not None:
        int_headers = [str(h).strip() if h else '' for h in int_rows[int_header_idx]]
        ic = {}
        for i, h in enumerate(int_headers):
            ic[h] = i

        for row in int_rows[int_header_idx + 1:]:
            if not row[0]:
                continue
            interaction = {
                'disruptor_id': str(row[ic['disruptor_id']] or '').strip(),
                'defender_id': str(row[ic['defender_id']] or '').strip(),
                'net_delta': to_num(row[ic['net_delta']]),
                'outcome_winner': str(row[ic['outcome_winner']] or '').strip(),
                'mechanism': str(row[ic['mechanism']] or '').strip(),
                'notes': str(row[ic['notes']] or '').strip(),
            }
            interactions.append(interaction)

        print(f'\nParsed {len(interactions)} interactions:')
        for ix in interactions:
            print(f"  {ix['disruptor_id']:8s} vs {ix['defender_id']:10s} | net_delta={ix['net_delta']:+.0f} | winner={ix['outcome_winner']}")

    # --- Output as JS global ---
    config = {
        'weapons': weapons,
        'interactions': interactions
    }

    json_str = json.dumps(config, indent=2, ensure_ascii=False)

    js_content = (
        '// Auto-generated from effect_matrix_v4.xlsx by scripts/parse_effect_matrix.py\n'
        '// Do not edit manually. Re-run: python scripts/parse_effect_matrix.py\n'
        '// ================================================================\n'
        f'const WEAPONS_CONFIG = {json_str};\n'
    )

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        f.write(js_content)

    print(f'\nWritten to {OUTPUT_PATH}')
    print(f'  {len(weapons)} weapons, {len(interactions)} interactions')


if __name__ == '__main__':
    main()
